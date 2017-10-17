IP_ADDRESS_COLUMN = 'A'
MODEL_VERSION_COLUMN='B'
IDRAC_VERSION_COLUMN='C'
SERVICE_TAG_COLUMN = 'D'
MEMORY_COLUMN = 'E'
MEMORY_SLOT_COLUMN = 'F'
CPU_COLUMN = 'G'
CPU_THREAD_COLUMN = 'H'
NIC_COLUMN = 'I'
PORT_COLUMN = 'J'
DRIVE_COLUMN = 'K'
DRIVE_SIZE_COLUMN = 'L'
MEMORY_INFO_COLUMN = 'M'
CPU_INFO_COLUMN = 'N'
NIC_INFO_COLUMN = 'O'
HDD_DRIVE_INFO_COLUMN = 'P'
SSD_DRIVE_INFO_COLUMN = 'Q'

USER_COLUMN = 'B'
PASSWORD_COLUMN = 'C'

try:
    import argparse
    import netaddr
    import Queue
    import time
    import getpass
    import string
    import paramiko
    import re
    from multiprocessing import Process
    import threading
    from threading import *
    # import subprocess
    import math
    from math import *
    # from subprocess import *
    import os
    import json
  #  import pandas as pd
  #  from pandas.io.json import json_normalize
    from datetime import datetime
    from ConfigParser import *
    from ConfigParser import (ConfigParser, MissingSectionHeaderError,
                              ParsingError, DEFAULTSECT)
    import ConfigParser
    import collections
    from collections import OrderedDict, Counter
    from openpyxl import *
    from openpyxl.styles import *
except ImportError as error:
    print "You don't have module \"{0}\" installed.".format(error.message[16:])
    exit(1)

# function for creating a xlsx with ip, user and password field


def createXlsxForIdracs(ip_set, user, pas):
    user = str(user)
    password = str(pas)
    wb = Workbook()
    ws = wb.active
    ws.title = "IdracInfo"
    columns_heading = ["IP", "User", "Pass"]
    columns_alphabat = ['A', 'B', 'C']
    yellowFill = PatternFill(start_color='FFFFFF00',
                             end_color='FFFFFF00', fill_type='solid')
    i = 0
    while(i < len(columns_heading)):
        ws[columns_alphabat[i] + str(1)].fill = yellowFill
        ws[columns_alphabat[i] + str(1)] = columns_heading[i]
        i += 1

    index = 2
    for ip in ip_set:
        str_index = str(index)
        ws[IP_ADDRESS_COLUMN + str_index].value = str(ip)
        ws[USER_COLUMN + str_index].value = user
        ws[PASSWORD_COLUMN + str_index].value = password
        index += 1
    name = 'IdracsList' + '.xlsx'
    wb.save(name)
    return index - 2, name

# fucntion for creating a xlsx by giving a number of idracs along with the three byte ip address and last byte


def createXlsxForIdrac(no_of_idracs, address, start):
    ip = address
    user = "root"
    password = "calvin"
    wb = Workbook()
    ws = wb.active
    ws.title = "IdracInfo"
    columns_heading = ["IP", "User", "Pass"]
    columns_alphabat = ['A', 'B', 'C']
    yellowFill = PatternFill(start_color='FFFFFF00',
                             end_color='FFFFFF00', fill_type='solid')
    i = 0
    while(i < len(columns_heading)):
        ws[columns_alphabat[i] + str(1)].fill = yellowFill
        ws[columns_alphabat[i] + str(1)] = columns_heading[i]
        i += 1
    index = 2

    for idrac in range(no_of_idracs):
        str_index = str(index)
        ws[IP_ADDRESS_COLUMN + str_index].value = ip + "." + str(start)
        ws[USER_COLUMN + str_index].value = user
        ws[PASSWORD_COLUMN + str_index].value = password
        start += 1
        index += 1

    name = 'IdracsList' + '.xlsx'
    wb.save(name)
    return name

# function for creating xlsx using json file


def CreateXLS(filename):
    with open(filename) as data_file:
        data = json.load(data_file)
    wb = Workbook()
    ws = wb.active
    # Column names
    columns_heading = ["IP","Model","Firmware Version", "Service Tag", "Total Memory", "Memory Slots", "Total CPU", "Total threads", "Total Nics",
                       "Total Ports", "Total Drives", "Total Drives Size", "Memory Info", "CPU Info", "NIC Info", "HDD Drives Info", "SSD Drives Info"]
    # column names correspond to excel
    columns_alphabat = ['A', 'B', 'C', 'D', 'E',
                        'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O','P','Q']
    i = 0
    # color
    yellowFill = PatternFill(start_color='FFFFFF00',
                             end_color='FFFFFF00', fill_type='solid')
    # set column names with color
    while(i < len(columns_heading)):
        ws[columns_alphabat[i] + str(1)].fill = yellowFill
        ws[columns_alphabat[i] + str(1)] = columns_heading[i]

#        ws.freeze_panes = ws[columns_alphabat[i] + str(1)]
        i += 1
    # freeze the top row
    ws.freeze_panes = 'A2'
    # Fill the data into the columns
    index = 2
    for idrac in range(len((data))):
        str_index = str(index)

        ws[IP_ADDRESS_COLUMN + str_index].value = data[idrac]['ip']
        ws[MODEL_VERSION_COLUMN + str_index].value = data[idrac]['model']
        ws[IDRAC_VERSION_COLUMN + str_index].value = data[idrac]['idrac_version']
        ws[SERVICE_TAG_COLUMN + str_index].value = data[idrac]['service_tag']
        ws[MEMORY_COLUMN +
            str_index].value = data[idrac]['total_memory'].split("/")[1]
        ws[MEMORY_SLOT_COLUMN + str_index].value = data[idrac]['memory_slots']
        ws[CPU_COLUMN + str_index].value = data[idrac]['cpu_slots']
        ws[CPU_THREAD_COLUMN + str_index].value = data[idrac]['cpu_threads']
        ws[NIC_COLUMN + str_index].value = data[idrac]['nic_slots']
        ws[PORT_COLUMN + str_index].value = data[idrac]['ports']

        drives = []
  #      print data[idrac]['drives_type']
        # Get to knmow the types of drives i.e. HDD and SSD and save it into a list
        for ind in range(len(data[idrac]['drives_type'])):
            for key, value in data[idrac]['drives_type'].iteritems():
                if key not in drives:
                    drives.append(key)
        # print drives
        tot_driv = 0
        # Calculate totaal drives using type (HDD or SSD)
        for ind in range(len(drives)):
           #         print data[idrac]['drives_type'][drives[ind]]
            tot_driv += int(data[idrac]['drives_type'][drives[ind]])
   #     print drives, tot_driv
        ws[DRIVE_COLUMN + str_index].value = tot_driv

        dumy_str = ''
        dumy_mem_model_list = []
        dumy_mem_speed_list = []
        mem_model_dict = dict()
        # fill the memory model list,speed list and fill a dictionary model vs speed
        for mem_slot in range(ws[MEMORY_SLOT_COLUMN + str_index].value):
            dumy_mem_model_list.append(
                data[idrac]['memory_info'][mem_slot]['model'])
            dumy_mem_speed_list.append(
                data[idrac]['memory_info'][mem_slot]['speed'])
            if data[idrac]['memory_info'][mem_slot]['model'] not in mem_model_dict:
                mem_model_dict[data[idrac]['memory_info'][mem_slot]
                               ['model']] = data[idrac]['memory_info'][mem_slot]['speed']
        # count the common memory model
        mem_model_set = Counter(dumy_mem_model_list)
  #      mem_model_set = collections.Counter(dumy_mem_model_list)
     #   print dumy_mem_model_list, mem_model_set, mem_model_dict
        # fill the memory info (model and speed)
        for mem in range(len(mem_model_set)):
            dumy_str += str(mem_model_set.keys()[mem]) + " x " + "Count: " + str(mem_model_set.values()[
                mem]) + " Speed: " + mem_model_dict[str(mem_model_set.keys()[mem])] + "\n"
       #     print dumy_str
        ws[MEMORY_INFO_COLUMN + str_index].value = dumy_str
        # align the values fr the xls
        ws[MEMORY_INFO_COLUMN + str_index].alignment = Alignment(wrapText=True)

        dumy_str = ''
        for cpu_slot in range(ws[CPU_COLUMN + str_index].value):
            dumy_str += data[idrac]['cpu_info'][cpu_slot]['model'] + "\t" + \
                " Processors: " + \
                data[idrac]['cpu_info'][cpu_slot]['no_of_processors'] + "\n"
        ws[CPU_INFO_COLUMN + str_index].value = dumy_str
        ws[CPU_INFO_COLUMN + str_index].alignment = Alignment(wrapText=True)

        dumy_str = ''
        for nic_slot in range(ws[NIC_COLUMN + str_index].value):
            #            for nested_index in range(len(data[idrac]['nic_info'][data[idrac]['nic_slots_name'][nic_slot]])):
            dumy_str += data[idrac]['nic_slots_name'][nic_slot] + \
                "\t" + " Type: " + data[idrac]['nic_info'][data[idrac]
                                                           ['nic_slots_name'][nic_slot]][0]['decription'] + "\n"
        ws[NIC_INFO_COLUMN + str_index].value = dumy_str
        ws[NIC_INFO_COLUMN + str_index].alignment = Alignment(wrapText=True)
        typ = 'GB'
        driv_type_dict = dict()

        drives_size_list = [list()
                            for x in range(len(data[idrac]['drives_type']))]
        drives_info_list = [list()
                            for x in range(len(data[idrac]['drives_type']))]

#	for ind in range(len(data[idrac]['drives_type'])):
 #           drives.append(data[idrac]['drives_type'][ind])
        total_mem = 0
        # Using drives type (HDD/SSD), find and count common drives i.e. 800 GB x 4 drives and add them into the excel sheet
        for driv_type in range(len(data[idrac]['drives_type'])):
            dumy_str = ''
            driv_type_dict = {}
            # Adjust the size of hardrives i.e write 586 GB as 600, 1024 GB as 1TB and make a dictionary size vs number (600 GB x 4)
            for driv_slot in range(data[idrac]['drives_type'][drives[driv_type]]):
                size_val = float(data[idrac]['drives_info'][drives[driv_type]]
                                 [driv_slot]['total_size_in_giga_bytes'].split(" ")[0])
                if 150 < size_val < 200:
                    size_val = 200
                    typ = " GB"
                elif 200 < size_val < 250:
                    size_val = 250
                    typ = " GB"
                elif 250 < size_val < 300:
                    size_val = 300
                    typ = " GB"
                elif 400 < size_val < 500:
                    size_val = 500
                    typ = " GB"
                elif 500 < size_val < 600:
                    size_val = 600
                    typ = " GB"
                elif size_val > 1024:
                    size_val = round(size_val / float(1024), 2)
                    typ = " TB"
                else:
                    size_val = size_val
                    typ = " GB"
               # print size_val

                drives_size_list[driv_type].append(size_val)
                # fill the dictionary (Size vs type "HDD or SSD")
                if str(size_val) not in driv_type_dict:
                    driv_type_dict[str(size_val)] = typ
            # Count the common drives
            drive_set = collections.Counter(drives_size_list[driv_type])
        #    print drive_set, driv_type_dict
            # calculate total drive size and fill the drive info column
            for drive in range(len(drive_set)):
                if "TB" in driv_type_dict[str(drive_set.keys()[drive])]:
                    drive_set.keys()[drive] = round(drive_set.keys()[drive], 2)
                    total_mem += (drive_set.keys()
                                  [drive]) * (drive_set.values()[drive]) * 1024
                elif "GB" in driv_type_dict[str(drive_set.keys()[drive])]:
                    # print type(drive_set.keys()[drive]), type(drive_set.values()[drive]), total_mem
                    total_mem += (drive_set.keys()
                                  [drive]) * (drive_set.values()[drive])
                dumy_str += str(drive_set.keys()[drive]) + driv_type_dict[str(drive_set.keys()[
                    drive])] + " x " + "Count: " + str(drive_set.values()[drive]) + "\n"
            # Save Drives Size and No of drives into list seperatey for HDD or SSD
            drives_info_list[driv_type] = dumy_str
     #   print total_mem
        # adjust total memory
        if total_mem > 1024:
            total_mem /= float(1024)
            total_mem = round(total_mem)
            total_mem = str(total_mem) + " TB"
        else:
            total_mem = str(total_mem) + " GB"
   #     print drives_info_list, total_mem
        # fill the HDD or SSD column of xlxs
        for ind in range(len(drives)):
            if "HDD" in drives[ind]:
                ws[HDD_DRIVE_INFO_COLUMN + str_index].value = drives_info_list[ind]
                ws[HDD_DRIVE_INFO_COLUMN +
                    str_index].alignment = Alignment(wrapText=True)
            elif "SSD" in drives[ind]:
                ws[SSD_DRIVE_INFO_COLUMN + str_index].value = drives_info_list[ind]
                ws[SSD_DRIVE_INFO_COLUMN +
                    str_index].alignment = Alignment(wrapText=True)

        ws[DRIVE_SIZE_COLUMN +
            str_index].value = (total_mem)
        index += 1
    # set the width of columns and save xlsx
    dims = {}
    max_siz = 0
    for row in ws.rows:
        for cell in row:
            if cell.value:
                if "\n" in str(cell.value) and len(str(cell.value)) > 100:
                    #    print str(cell.value),len(str(cell.value))
                    dumy = str(cell.value).split("\n")
                    for i in range(len(dumy)):
                        siz = len(dumy[i])
                        if siz > max_siz:
                            max_siz = siz
                    dims[cell.column] = max(
                        (dims.get(cell.column, 0), max_siz))
                else:
                    dims[cell.column] = max(
                        (dims.get(cell.column, 0), len(str(cell.value))))
    for col, value in dims.items():
        ws.column_dimensions[col].width = value
    timestr = time.strftime("%Y%m%d_%H%M%S")
    name = 'Output' + timestr + '.xlsx'
    wb.save(name)

# function will ssh into an idrac and saves the output into the HwList


def process_idrac(idrac_info, HwList, idracNo, logfptr):
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    try:
        ssh.connect(idrac_info[0], username=idrac_info[1],
                    password=idrac_info[2])
        stdin, stdout, stderr = ssh.exec_command("racadm hwinventory")
        HwList[idracNo] = stdout.readlines()
        ssh.close()
    except Exception as e:
        time.sleep(2)
        logfptr.write("ssh.connect not work for Idrac: %s %s %s %s\n" %
                      (idrac_info[0], idrac_info[1], idrac_info[2], e))
#        print "ssh.connect not work for Idrac: ", idrac_info[0], idrac_info[1], idrac_info[2], e
    return

# funtion creates threads for idracs for ssh and returns their output into a list


def sshAndStoreData(IdracList, logfptr):

    HwInventoryList = [0] * len(IdracList)
    # print len(HwInventoryList)
    threads = []
    for idracNo in range(0, len(IdracList)):

        nam = "Idrac" + str(idracNo)
        #Process(target=process_idrac, args=(HwInventoryList[idracNo], IdracList[idracNo][0], IdracList[idracNo][1], IdracList[idracNo][2],)).start()
        thread = threading.Thread(
            name=nam, target=process_idrac, args=(IdracList[idracNo], HwInventoryList, idracNo, logfptr,))
        thread.start()
        threads.append(thread)
    for x in threads:
        x.join()
    return HwInventoryList


# idrac default usr and password
IDRAC_FACTORY_ADMIN_USER_CREDENTIALS = {
    'user_name': 'root',
    'password': 'calvin'}


def ip_address_from_address(address):
    try:
        ip_address = netaddr.IPAddress(address)
     #   print ip_address
    except ValueError as e:
        # address contains a CIDR prefix, netmask, or hostmask.
        e.message = ('invalid IP address: %(address)s (detected CIDR, netmask,'
                     ' hostmask, or subnet)') % {'address': address}
        raise
    except netaddr.AddrFormatError as e:
        # address is not an IP address represented in an accepted string
        # format.
        e.message = ("invalid IP address: '%(address)s' (failed to detect a"
                     " valid IP address)") % {'address': address}
        raise

    if ip_address.version == 4:
        return ip_address
    else:
        raise NotSupported(
            ('invalid IP address: %(address)s (Internet Protocol version'
             ' %(version)s is not supported)') % {
                'address': address,
                'version': ip_address.version})


def ip_set_from_address_range(start, end):
    try:
        start_ip_address = ip_address_from_address(start)
        end_ip_address = ip_address_from_address(end)
       # print start_ip_address, end_ip_address
    except (NotSupported, ValueError) as e:
        raise ValueError(
            ('invalid IP range: %(start)s-%(end)s (%(message)s)') %
            {
                'start': start,
                'end': end,
                'message': e.message})
    except netaddr.AddrFormatError as e:
        raise ValueError(
            ("invalid IP range: '%(start)s-%(end)s' (%(message)s)") %
            {
                'start': start,
                'end': end,
                'message': e.message})

    if start_ip_address > end_ip_address:
        raise ValueError(
            ('invalid IP range: %(start)s-%(end)s (lower bound IP greater than'
             ' upper bound)') %
            {
                'start': start,
                'end': end})

    ip_range = netaddr.IPRange(start_ip_address, end_ip_address)

    return netaddr.IPSet(ip_range)


def ip_set_from_address(address):
    ip_set = netaddr.IPSet()
    # print address
    try:
        ip_address = ip_address_from_address(address)
        ip_set.add(ip_address)
    except ValueError:
        ip_network = ip_network_from_address(address)
        ip_set.update(ip_network.iter_hosts())

    return ip_set


def parse_idrac_arguments(idrac_list):
    ip_set = netaddr.IPSet()
    for idrac in idrac_list:
        ip_set = ip_set.union(ip_set_from_idrac(idrac))
    return ip_set


def ip_set_from_idrac(idrac):
    range_bounds = idrac.split('-')
    if len(range_bounds) == 2:
        start, end = range_bounds
        ip_set = ip_set_from_address_range(start, end)
    elif len(range_bounds) == 1:
        ip_set = ip_set_from_address(range_bounds[0])
    else:
        # String contains more than one (1) dash.
        raise ValueError(
            ('invalid IP range: %(idrac)s (contains more than one hyphen)') % {
                'idrac': idrac})

    return ip_set

#fucntion used to extract the key value of a section
def get_key_value_from_section(key_value_dict,desired_key):
    missing="Value Missing"
 #   print desired_key.lower()
    if desired_key.lower() in key_value_dict:
      #  print desired_key.lower(), key_value_dict[desired_key.lower()]
        return str(key_value_dict[desired_key.lower()])
    else:
        return missing
            

def extract_mem(Config):
    Mem_section_list, mem_size_list, mem_speed_list, mem_descp_list, mem_serial_list, mem_model_list, mem_status_list = ([
    ]for i in range(0, 7))
    for s in Config.sections():
        if "InstanceID: DIMM.Socket" in s:
            Mem_section_list += [s]
 #       Mem_section_list = [s for s, s in enumerate(
  #          Config.sections()) if "InstanceID: DIMM.Socket" in s]
        # Extract key, values from memory sections
    for mem_sec in Mem_section_list:
        key_valu_dict = dict(Config.items(mem_sec))
  #      print key_valu_dict
        mem_size_list += [get_key_value_from_section(key_valu_dict, 'Size')]
        mem_speed_list += [get_key_value_from_section(key_valu_dict,
                                      'CurrentOperatingSpeed')]
        mem_descp_list += [get_key_value_from_section(key_valu_dict, 'Manufacturer')]
        mem_serial_list += [get_key_value_from_section(key_valu_dict, 'SerialNumber')]
        mem_model_list += [get_key_value_from_section(key_valu_dict, 'Model')]
        mem_status_list += [get_key_value_from_section(key_valu_dict, 'PrimaryStatus')]
        # Check the types and sizes of memory, convert it into MB, and store them into a list
    for index in range(0, len(mem_size_list)):
        if "MB" in mem_size_list[index]:
            value = mem_size_list[index].replace("MB", "")
            mem_size_list[index] = value
        elif "PB" in mem_size_list[index]:
            value = mem_size_list[index].replace("PB", "")
            value *= math.pow(1024, 3)
            mem_size_list[index] = value
        elif "TB" in mem_size_list[index]:
            value = mem_size_list[index].replace("TB", "")
            value *= math.pow(1024, 2)
            mem_size_list[index] = value
        elif "GB" in mem_size_list[index]:
            value = mem_size_list[index].replace("GB", "")
            value *= math.pow(1024, 1)
            mem_size_list[index] = value
        elif "KB" in mem_size_list[index]:
            value = mem_size_list[index].replace("KB", "")
            value /= 1024
            mem_size_list[index] = value
        # total memory slots
    mem_slots = len(mem_size_list)
    # list of dictionaries for storing features of each memory slot
    mem_info_dictionaries = [dict() for x in range(mem_slots)]
    mem_info_list = []

    # Store memory key value individually intothe list of dictionaries
    for i in range(mem_slots):
        mem_info_dictionaries[i]['manufacturer'] = mem_descp_list[i]
        mem_info_dictionaries[i]['speed'] = mem_speed_list[i]
        mem_info_dictionaries[i]['model'] = mem_model_list[i]
        mem_info_dictionaries[i]['serial'] = mem_serial_list[i]
        mem_info_dictionaries[i]['status'] = mem_status_list[i]
    # Store list of dictionaries into a list
    mem_info_list = mem_info_dictionaries
    total_mem = str(sum([int(x) for x in mem_size_list]))
#    print total_mem, mem_info_list, mem_slots
    return total_mem, mem_info_list, mem_slots


def extract_cpu(Config):

    # Lists variable initialization
    Cpu_section_list, cpu_processor_list, cpu_family_list, cpu_manufac_list, cpu_curr_clock_list, cpu_model_list = ([
    ]for i in range(6))
    cpu_prim_status_list, cpu_virt_list, cpu_voltag_list, cpu_enabled_thread_list, cpu_max_clock_speed_list = ([
    ]for i in range(5))
    cpu_ex_bus_clock_speed_list, cpu_hyper_thread_list, cpu_status_list = (
        []for i in range(3))

    # Store cpu sections into the list
    for s in Config.sections():
        if "InstanceID: CPU.Socket" in s:
            Cpu_section_list += [s]
    for cpu_sec in Cpu_section_list:
        key_valu_dict = dict(Config.items(cpu_sec))
        
        cpu_processor_list += [get_key_value_from_section(key_valu_dict,
                                          'NumberOfProcessorCores')]
        cpu_family_list += [get_key_value_from_section(key_valu_dict, 'CPUFamily')]
        cpu_manufac_list += [get_key_value_from_section(key_valu_dict, 'Manufacturer')]
        cpu_curr_clock_list += [get_key_value_from_section(key_valu_dict,
                                           'CurrentClockSpeed')]
        cpu_model_list += [get_key_value_from_section(key_valu_dict, 'Model')]
        cpu_prim_status_list += [get_key_value_from_section(key_valu_dict, 'PrimaryStatus')]
        cpu_virt_list += [get_key_value_from_section(key_valu_dict,
                                     'VirtualizationTechnologyEnabled')]
        cpu_voltag_list += [get_key_value_from_section(key_valu_dict, 'Voltage')]
        cpu_enabled_thread_list += [get_key_value_from_section(key_valu_dict, 'NumberOfEnabledThreads')]
        cpu_max_clock_speed_list += [
            get_key_value_from_section(key_valu_dict, 'MaxClockSpeed')]
        cpu_ex_bus_clock_speed_list += [get_key_value_from_section(key_valu_dict, 'ExternalBusCLockSpeed')]
        cpu_hyper_thread_list += [get_key_value_from_section(key_valu_dict,
                                             'HyperThreadingEnabled')]
        cpu_status_list += [get_key_value_from_section(key_valu_dict, 'CPUStatus')]
        
    cpu_slots = len(cpu_processor_list)
    cpu_info_dictionaries = [dict() for x in range(cpu_slots)]
    cpu_info_list = []
    total_threads = 0
    # Store cpu key value individually into the list of dictionaries
    for i in range(cpu_slots):
        cpu_info_dictionaries[i]['no_of_processors'] = cpu_processor_list[i]
        cpu_info_dictionaries[i]['cpu_family'] = cpu_family_list[i]
        cpu_info_dictionaries[i]['manufacturer'] = cpu_manufac_list[i]
        cpu_info_dictionaries[i]['current_clock_speed'] = cpu_curr_clock_list[i]
        cpu_info_dictionaries[i]['model'] = cpu_model_list[i]
        cpu_info_dictionaries[i]['primary_status'] = cpu_prim_status_list[i]
        cpu_info_dictionaries[i]['virtualization_technology_enabled'] = cpu_virt_list[i]
        cpu_info_dictionaries[i]['voltage'] = cpu_voltag_list[i]
        cpu_info_dictionaries[i]['no_of_enabled_thread'] = cpu_enabled_thread_list[i]
        cpu_info_dictionaries[i]['max_clock_Speed'] = cpu_max_clock_speed_list[i]
        cpu_info_dictionaries[i]['external_bus_clock_speed'] = cpu_ex_bus_clock_speed_list[i]
        cpu_info_dictionaries[i]['hyper_threading_enabled'] = cpu_hyper_thread_list[i]
        total_threads += int(cpu_enabled_thread_list[i])
        cpu_info_dictionaries[i]['cpu_status'] = cpu_status_list[i]
        # Store list of dictionaries into a list
    cpu_info_list = cpu_info_dictionaries
   # print cpu_slots, total_threads, cpu_info_list
    return cpu_slots, total_threads, cpu_info_list


def extract_nics(Config):

    # Lists variable initialization
    Nic_section_list, nic_manufac_list, nic_descrp_list, nic_dev_descrp_list, nic_fqdd_list = ([
    ]for i in range(5))
    nic_pci_sub_device_id_list, nic_pci_sub_vendor_id_list, nic_pci_device_id_list = (
        []for i in range(3))
    nic_pci_vendor_id_list, nic_bus_no_list, nic_slot_typ_list, nic_data_bus_width_list = ([
    ]for i in range(4))
    nic_fun_no_list, nic_dev_no_list = ([]for i in range(2))
    # Store nic sections into the list
    for s in Config.sections():
        if "InstanceID: NIC" in s:
            Nic_section_list += [s]
    Nic_set = set()
    port_slots = len(Nic_section_list)
    for section in Nic_section_list:
        nic = section.split(" ")[1].split(".")[0] + "." + section.split(" ")[1].split(
            ".")[1] + "." + section.split(" ")[1].split(".")[2].split("-")[0]
        if nic not in Nic_set:
            Nic_set.add(nic)
    nic_slots = len(Nic_set)
    # Using the set of nics, make a dictionary which corresponds to the ports each nic contains
    Nic_set_dic = dict()
    for nic in Nic_set:
        for nic_sec in Nic_section_list:
            if nic in nic_sec:
                if nic not in Nic_set_dic:
                    Nic_set_dic[nic] = list()
                    Nic_set_dic[nic].append(nic_sec)
                else:
                    Nic_set_dic[nic].append(nic_sec)

        # Extract key, values from nic sections
    for nic_sec in Nic_section_list:
        key_valu_dict  = dict(Config.items(nic_sec))
        nic_manufac_list += [get_key_value_from_section(key_valu_dict, 'Manufacturer')]
        nic_descrp_list += [get_key_value_from_section(key_valu_dict, 'Description')]
        nic_dev_descrp_list += [get_key_value_from_section(key_valu_dict,
                                           'DeviceDescription')]
        nic_fqdd_list += [get_key_value_from_section(key_valu_dict, 'FQDD')]
        nic_pci_sub_device_id_list += [
            get_key_value_from_section(key_valu_dict, 'PCISubDeviceID')]
        nic_pci_sub_vendor_id_list += [
            get_key_value_from_section(key_valu_dict, 'PCISubVendorID')]
        nic_pci_device_id_list += [get_key_value_from_section(key_valu_dict, 'PCIDeviceID')]
        nic_pci_vendor_id_list += [get_key_value_from_section(key_valu_dict, 'PCIVendorID')]
        nic_bus_no_list += [get_key_value_from_section(key_valu_dict, 'BusNumber')]
        nic_slot_typ_list += [get_key_value_from_section(key_valu_dict, 'SlotType')]
        nic_data_bus_width_list += [
            get_key_value_from_section(key_valu_dict, 'DataBusWidth')]
        nic_fun_no_list += [get_key_value_from_section(key_valu_dict, 'FunctionNumber')]
        nic_dev_no_list += [get_key_value_from_section(key_valu_dict, 'DeviceNumber')]

    nic_info_dictionaries = [dict() for x in range(port_slots)]
    # Store nic key value individually into the list of dictionaries
    for i in range(port_slots):
        nic_info_dictionaries[i]['manufacturer'] = nic_manufac_list[i]
        nic_info_dictionaries[i]['decription'] = nic_descrp_list[i]
        nic_info_dictionaries[i]['device_description'] = nic_dev_descrp_list[i]
        nic_info_dictionaries[i]['fqdd'] = nic_fqdd_list[i]
        nic_info_dictionaries[i]['pci_subdevice_id'] = nic_pci_sub_device_id_list[i]
        nic_info_dictionaries[i]['pci_Subvendor_id'] = nic_pci_sub_vendor_id_list[i]
        nic_info_dictionaries[i]['pci_Dev_id'] = nic_pci_device_id_list[i]
        nic_info_dictionaries[i]['pci_vendor_id'] = nic_pci_vendor_id_list[i]
        nic_info_dictionaries[i]['bus_number'] = nic_bus_no_list[i]
        nic_info_dictionaries[i]['slot_type'] = nic_slot_typ_list[i]
        nic_info_dictionaries[i]['data_bus_width'] = nic_data_bus_width_list[i]
        nic_info_dictionaries[i]['function_number'] = nic_fun_no_list[i]
        nic_info_dictionaries[i]['dev_number'] = nic_dev_no_list[i]
        # Store keys and values of nic in dictionary using a nic set vs port formate
    nic_port_dictionaries = dict()
    nic_name_list = []
    for nic in Nic_set:
        nic_name_list.append(nic)
        dumy_dict = [dict() for x in range(len(Nic_set_dic[nic]))]
        ind = 0
        for nics in Nic_set_dic[nic]:
            for j in range(port_slots):
                if nic_info_dictionaries[j]['fqdd'] in nics:
                    dumy_dict[ind] = nic_info_dictionaries[j]
                    ind += 1
        nic_port_dictionaries[nic] = dumy_dict
#    print nic_slots, nic_name_list, port_slots, nic_port_dictionaries
    return nic_slots, nic_name_list, port_slots, nic_port_dictionaries


def extract_hard_drives(Config):
    # Lists variable initialization
    Drive_section_list, driv_dev_list, driv_dev_descrp_list, driv_ppid_list, temp_size_list, = ([
    ]for i in range(5))
    driv_sas_address_list, driv_max_cap_speed_list, driv_used_siz_byte_list, driv_media_typ_list = ([
    ]for i in range(4))
    driv_blck_siz_byte_list, driv_bus_protocol_list, driv_serial_no_list, driv_manufacturer_list = ([
    ]for i in range(4))
    driv_fqdd_list, driv_slot_list, driv_raid_status_list, driv_predictiv_fail_status_list, driv_free_siz_byte_list, driv_total_siz_list = ([
    ]for i in range(6))
    # Store drive sections into the list
    for s in Config.sections():
        if "InstanceID: Disk" in s:
            if ("Virtual" in s) or ("FlashCard" in s):
                continue;
            else:
                Drive_section_list += [s]
                
    drive_list_len = len(Drive_section_list)
    
    # Extract key, values from drives sections
    check_ssd = 0
    check_hdd = 0
    for dev_sec in Drive_section_list:
        key_valu_dict = dict(Config.items(dev_sec))
        driv_dev_list += [get_key_value_from_section(key_valu_dict, 'Device Type')]
        driv_dev_descrp_list += [get_key_value_from_section(key_valu_dict,
                                            'DeviceDescription')]
        driv_ppid_list += [get_key_value_from_section(key_valu_dict, 'PPID')]
        driv_sas_address_list += [get_key_value_from_section(key_valu_dict, 'SASAddress')]
        driv_max_cap_speed_list += [
            get_key_value_from_section(key_valu_dict, 'MaxCapableSpeed')]
        driv_used_siz_byte_list += [
            get_key_value_from_section(key_valu_dict, 'UsedSizeInBytes')]
        driv_free_siz_byte_list += [
            get_key_value_from_section(key_valu_dict, 'FreeSizeInBytes')]
        total_hard_size = int(get_key_value_from_section(key_valu_dict, 'UsedSizeInBytes').split(" ")[0]) + int(
            get_key_value_from_section(key_valu_dict, 'FreeSizeInBytes').split(" ")[0])
        total_hard_size /= math.pow(1024, 3)
#	    print total_hard_size
        driv_total_siz_list.append(total_hard_size)
        driv_media_typ_list += [get_key_value_from_section(key_valu_dict, 'MediaType')]
        # check and count the total ssd or hdd
        if ("Solid State Drive" in get_key_value_from_section(key_valu_dict, 'MediaType')):
            check_ssd += 1
        if ("HDD" in get_key_value_from_section(key_valu_dict, 'MediaType') or ("Magnetic Drive" in get_key_value_from_section(key_valu_dict, 'MediaType'))):
            check_hdd += 1
        driv_blck_siz_byte_list += [
            get_key_value_from_section(key_valu_dict, 'BlockSizeInBytes')]
        driv_bus_protocol_list += [get_key_value_from_section(key_valu_dict, 'BusProtocol')]
        driv_serial_no_list += [get_key_value_from_section(key_valu_dict, 'SerialNumber')]
        driv_manufacturer_list += [
            get_key_value_from_section(key_valu_dict, 'Manufacturer')]
        driv_fqdd_list += [get_key_value_from_section(key_valu_dict, 'FQDD')]
        driv_slot_list += [get_key_value_from_section(key_valu_dict, 'Slot')]
        driv_raid_status_list += [get_key_value_from_section(key_valu_dict, 'RaidStatus')]
        driv_predictiv_fail_status_list += [
            get_key_value_from_section(key_valu_dict, 'PredictiveFailureState')]
        
#        print driv_total_siz_list
    total_size = (sum([int(x) for x in driv_total_siz_list]))
    driv_slots = len(Drive_section_list)
    
    # list of dictionaries for hdd and ssd for storing each slot information seperately
    driv_info_hdd_dictionaries = [dict() for x in range(check_hdd)]
    driv_info_ssd_dictionaries = [dict() for x in range(check_ssd)]
    driv_info_list = []
    # Store hardrive key value individually into the list of dictionaries
    hdd_index = 0
    ssd_index = 0
    for i in range(driv_slots):
        if "HDD" in driv_media_typ_list[i] or "Magnetic Drive" in driv_media_typ_list[i]:
            driv_info_hdd_dictionaries[hdd_index]['device_type'] = driv_dev_list[i]
            driv_info_hdd_dictionaries[hdd_index]['device_description'] = driv_dev_descrp_list[i]
            driv_info_hdd_dictionaries[hdd_index]['ppid'] = driv_ppid_list[i]
            driv_info_hdd_dictionaries[hdd_index]['sas_address'] = driv_sas_address_list[i]
            driv_info_hdd_dictionaries[hdd_index]['max_capable_speed'] = driv_max_cap_speed_list[i]
            driv_info_hdd_dictionaries[hdd_index]['used_size_in_bytes'] = driv_used_siz_byte_list[i]
            driv_info_hdd_dictionaries[hdd_index]['free_size_in_bytes'] = driv_free_siz_byte_list[i]
            driv_info_hdd_dictionaries[hdd_index]['total_size_in_giga_bytes'] = str(
                driv_total_siz_list[i]) + " GB"
#		    driv_info_hdd_dictionaries[hdd_index]['media_type'] = driv_media_typ_list[i]
            driv_info_hdd_dictionaries[hdd_index]['block_size_in_bytes'] = driv_blck_siz_byte_list[i]
            driv_info_hdd_dictionaries[hdd_index]['bus_protocol'] = driv_bus_protocol_list[i]
            driv_info_hdd_dictionaries[hdd_index]['serial_no'] = driv_serial_no_list[i]
            driv_info_hdd_dictionaries[hdd_index]['manufacturer'] = driv_manufacturer_list[i]
            driv_info_hdd_dictionaries[hdd_index]['fqdd'] = driv_fqdd_list[i]
            driv_info_hdd_dictionaries[hdd_index]['slot'] = driv_slot_list[i]
            driv_info_hdd_dictionaries[hdd_index]['raid_status'] = driv_raid_status_list[i]
            driv_info_hdd_dictionaries[hdd_index]['predictive_failure_state'] = driv_predictiv_fail_status_list[i]
            hdd_index += 1
        elif "Solid State Drive" in driv_media_typ_list[i]:
            driv_info_ssd_dictionaries[ssd_index]['device_type'] = driv_dev_list[i]
            driv_info_ssd_dictionaries[ssd_index]['device_description'] = driv_dev_descrp_list[i]
            driv_info_ssd_dictionaries[ssd_index]['ppid'] = driv_ppid_list[i]
            driv_info_ssd_dictionaries[ssd_index]['sas_address'] = driv_sas_address_list[i]
            driv_info_ssd_dictionaries[ssd_index]['max_capable_speed'] = driv_max_cap_speed_list[i]
            driv_info_ssd_dictionaries[ssd_index]['used_size_in_bytes'] = driv_used_siz_byte_list[i]
            driv_info_ssd_dictionaries[ssd_index]['free_size_in_bytes'] = driv_free_siz_byte_list[i]

            driv_info_ssd_dictionaries[ssd_index]['total_size_in_giga_bytes'] = str(
                driv_total_siz_list[i]) + " GB"
#		    driv_info_ssd_dictionaries[ssd_index]['media_type'] = driv_media_typ_list[i]
            driv_info_ssd_dictionaries[ssd_index]['block_size_in_bytes'] = driv_blck_siz_byte_list[i]
            driv_info_ssd_dictionaries[ssd_index]['bus_protocol'] = driv_bus_protocol_list[i]
            driv_info_ssd_dictionaries[ssd_index]['serial_no'] = driv_serial_no_list[i]
            driv_info_ssd_dictionaries[ssd_index]['manufacturer'] = driv_manufacturer_list[i]
            driv_info_ssd_dictionaries[ssd_index]['fqdd'] = driv_fqdd_list[i]
            driv_info_ssd_dictionaries[ssd_index]['slot'] = driv_slot_list[i]
            driv_info_ssd_dictionaries[ssd_index]['raid_status'] = driv_raid_status_list[i]
            driv_info_ssd_dictionaries[ssd_index]['predictive_failure_state'] = driv_predictiv_fail_status_list[i]
            ssd_index += 1

    driv_dict = dict()
    driv_type_dict = dict()

    if check_ssd > 0:
        # save ssd dictionaries features into a dictionary
        driv_dict['SSD'] = driv_info_ssd_dictionaries
        # save drives type vs total drive of that type "ssd" into a dictioanary
        driv_type_dict['SSD'] = ssd_index
    if check_hdd > 0:
        driv_dict['HDD'] = driv_info_hdd_dictionaries
        driv_type_dict['HDD'] = hdd_index
  #  print driv_type_dict, total_size, driv_dict
    return driv_type_dict, total_size, driv_dict


def main():

    # parse input argument
    parser = argparse.ArgumentParser(description='Idrac Inventory')
    parser.add_argument('-i','--idrac', nargs='+', help="""White space separated list of IP address specifications to scan for iDRACs. Each specification may be an IP address, range of IP addresses. A range's start and end IP addresses are separated by a hyphen.Only IPv4 addresses are supported.""")
    parser.add_argument(
        '-u', '--username', default=IDRAC_FACTORY_ADMIN_USER_CREDENTIALS['user_name'], help='username for accessing the iDRACs')
    parser.add_argument(
        '-p', '--password', default=IDRAC_FACTORY_ADMIN_USER_CREDENTIALS['password'], help='password for accessing the iDRACs')
    parser.add_argument(
        '-e', '--excel', default='no', help='excel sheet for all the idrac(s)')
    args = parser.parse_args()
    # get a list of ip address in a form of set
    ip_set = parse_idrac_arguments(idrac_list=args.idrac)
    # save ip, user and password into the xlsx and get no of idracs
    NoOfIdracs, IdracListXlsxFile = createXlsxForIdracs(
        ip_set, args.username, args.password)

    IdracList = []
    Idrac = []
    print "loading file"
    # load xlsx sheet
    wb = load_workbook(IdracListXlsxFile)
    ws = wb['IdracInfo']
    xls_ind = 2
    # fill the idrac list with ip, user and password from xlsx
    for i in range(NoOfIdracs):
        Idrac = []
        Idrac.append(ws[IP_ADDRESS_COLUMN + str(xls_ind)].value)
        Idrac.append(ws[USER_COLUMN + str(xls_ind)].value)
        Idrac.append(ws[PASSWORD_COLUMN + str(xls_ind)].value)
        IdracList.append(Idrac)
        xls_ind += 1
    # final list contains the ip, user and password
    print IdracList
    # no idrac
    if len(IdracList) == 0:
        print "IdracList is empty, terminating the application"
        exit()
    # file for logging
    Logfilename = "log" + ".txt"
    logfptr = open(Logfilename, "w+")

    # Using the stored information ssh into the server and get Hwinventory
    HwInventoryList = sshAndStoreData(IdracList, logfptr)

    TotalIdracsOutputs = len(HwInventoryList)
    # Final list of idracs
    iDRAC_list = []

    Outfilename = "Idrac" + ".json"

    continue_now = 0
    for IdracNo in range(0, TotalIdracsOutputs):
        #	print IdracNo
        continue_now = 0
        # list is empty although we tried to ssh into the idrac
        if not HwInventoryList[IdracNo]:
            time.sleep(1)
            logfptr.write("HwInventory list is empty for idrac: %s %s %s\n" % (
                IdracList[IdracNo][0], IdracList[IdracNo][1], IdracList[IdracNo][2]))
       #     print "HwInventory list is empty for idrac: ", IdracList[IdracNo][0],IdracList[IdracNo][1],IdracList[IdracNo][2]
            continue
        else:
            pass
        # Lists and dictionaries variable initialization

        idrac_dumy_dic = OrderedDict()

   #     time = time.strftime("%Y%m%d-%H%M%S")
        filename = "Idrac" + str(IdracNo) + ".ini"
        fptr = open(filename, "w+")

        # Read Hwinventory and write it into the file
        for line in HwInventoryList[IdracNo]:
            if "--" in line:
                continue
            elif "ERROR: Invalid subcommand specified" in line:
                logfptr.write("%s for %s %s %s\n" % (
                    line, IdracList[IdracNo][0], IdracList[IdracNo][1], IdracList[IdracNo][2]))
                continue_now = 1
                break
            elif "No more sessions are available" in line:
                logfptr.write("%s for %s %s %s\n" % (
                    line, IdracList[IdracNo][0], IdracList[IdracNo][1], IdracList[IdracNo][2]))
                continue_now = 1
                break
            fptr.write("%s\n" % line)
        fptr.close()
        # Racadm command did not work for those idracs so delete the ini file and continue
        if continue_now:
            os.remove(filename)
            continue

        # Read Sections from filename
        try:
            Config = ConfigParser.ConfigParser()
            Config.read(filename)
#	    print Config.sections()
        except (ConfigParser.MissingSectionHeaderError,
                ConfigParser.ParsingError):
            logfptr.write("No Section header for %s %s %s\n" % (
                IdracList[IdracNo][0], IdracList[IdracNo][1], IdracList[IdracNo][2]))
            continue

        total_mem, mem_info_list, mem_slots = extract_mem(Config)

        idrac_dumy_dic['ip'] = str(IdracList[IdracNo][0])

        idrac_dumy_dic['total_memory'] = total_mem + " MB" + \
            " / " + str(int(total_mem) / 1024) + " GB"
        idrac_dumy_dic['memory_slots'] = mem_slots
        idrac_dumy_dic['memory_info'] = mem_info_list

        cpu_slots, total_threads, cpu_info_list = extract_cpu(Config)
        idrac_dumy_dic['cpu_slots'] = cpu_slots
        idrac_dumy_dic['cpu_threads'] = total_threads
        idrac_dumy_dic['cpu_info'] = cpu_info_list

        nic_slots, nic_name_list, port_slots, nic_port_dictionaries = extract_nics(
            Config)
        idrac_dumy_dic['nic_slots'] = nic_slots
        idrac_dumy_dic['nic_slots_name'] = nic_name_list
        idrac_dumy_dic['ports'] = port_slots
        idrac_dumy_dic['nic_info'] = nic_port_dictionaries

        driv_type_dict, total_size, driv_dict = extract_hard_drives(Config)
        idrac_dumy_dic['drives_type'] = driv_type_dict
        idrac_dumy_dic['total_drives_size'] = str(total_size) + " GB"
        idrac_dumy_dic['drives_info'] = driv_dict

        Sys_section_list = []
        # Store sys sections into the list
        for s in Config.sections():
            if "InstanceID: System.Embedded" in s:
                idrac_dumy_dic['service_tag'] = Config.get(
                    str(s), 'ServiceTag')
                idrac_dumy_dic['model'] = Config.get(
                    str(s), 'Model').split(" ")[1]

            elif "InstanceID: iDRAC." in s:
                idrac_dumy_dic['idrac_version'] = Config.get(
                    str(s), 'FirmwareVersion')
              #  Sys_section_list += [s]

      #  idrac_dumy_dic['service_tag'] = Config.get(
       #     str(Sys_section_list[0]), 'ServiceTag')

        # Finally add all the nested dictionaries and list into a list (json formate)
        iDRAC_list.append(idrac_dumy_dic)

    logfptr.close()
    # Store the json formate into a file
    fptr = open(Outfilename, "w+")
    json.dump(iDRAC_list, fptr, indent=4,
              sort_keys=False, separators=(',', ':'))
    fptr.close()
    # Create Excel file from json
    if args.excel == 'yes':
        CreateXLS(Outfilename)


main()
